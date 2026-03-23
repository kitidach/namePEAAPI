#!/usr/bin/env python3
"""
PEA Data Extraction & Normalization Script
==========================================

Reads Peaname.xlsx Sheet1 and produces:
  - data/pea_master.json       (canonical normalized data)
  - data/validation_report.json (issues & warnings)

Key insight: Sheet1 contains ~3 repeated sections of the same offices 
(rows 2-120, 121-250, 251+). Each section shows the same offices with 
slightly different grouping context (different Q/R/S values). We take 
the FIRST occurrence of each unique P code as the canonical record.

Usage:
  python scripts/extract_pea_data.py

Requires: openpyxl
"""

import json
import re
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# ─── Paths ───────────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "data"
EXCEL_PATH = DATA_DIR / "Peaname.xlsx"
OUTPUT_PATH = DATA_DIR / "pea_master.json"
REPORT_PATH = DATA_DIR / "validation_report.json"


# ─── Normalization Utilities ─────────────────────────────────────────

def normalize_thai_text(text: str) -> str:
    """Normalize Thai text for search: strip, collapse spaces."""
    if not text:
        return ""
    text = str(text).strip()
    text = re.sub(r'\s+', ' ', text)
    return text


def generate_search_tokens(office: dict) -> list:
    """Generate normalized search tokens from all searchable fields."""
    fields = [
        office.get("code", ""),
        office.get("shortName", ""),
        office.get("fullName", ""),
    ]
    fields.extend(office.get("aliases", []))
    fields.extend([
        office.get("parent43Name", ""),
        office.get("parent17Name", ""),
        office.get("parent8Name", ""),
        office.get("regionName", ""),
    ])

    tokens = set()
    for f in fields:
        if not f:
            continue
        val = normalize_thai_text(str(f))
        if val:
            tokens.add(val)
            # Add version without dots for fuzzy Thai abbreviation search
            no_dots = val.replace(".", "").strip()
            if no_dots and no_dots != val:
                tokens.add(no_dots)
    return sorted(tokens)


def parse_index_and_name(h_value: str):
    """
    Parse col H like '1. กฟจ.อุดรธานี' -> (1, 'กฟจ.อุดรธานี')
    Also handles '0. อุดรธานี ฉ.1' for region row.
    """
    if not h_value:
        return None, None
    h_str = str(h_value).strip()
    match = re.match(r'^(\d+)\.\s*(.+)$', h_str)
    if match:
        return int(match.group(1)), match.group(2).strip()
    return None, h_str


def safe_str(val) -> str:
    """Convert value to string safely. Returns '' for None/#N/A."""
    if val is None:
        return ""
    s = str(val).strip()
    if s in ("#N/A", "#REF!", "#VALUE!", "#NAME?"):
        return ""
    return s


def safe_int(val):
    """Convert to int safely. Returns None for non-numeric."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    try:
        return int(str(val).strip())
    except (ValueError, TypeError):
        return None


# ─── Main Extraction ─────────────────────────────────────────────────

def extract():
    """Main extraction pipeline."""
    report = {
        "source": str(EXCEL_PATH),
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "warnings": [],
        "errors": [],
        "stats": {},
        "assumptions": [
            "P code (col P) is the unique primary key for offices",
            "Sheet1 has ~3 repeated sections with same offices in different grouping contexts",
            "We take the FIRST occurrence of each P code as the canonical record",
            "Rows with P='D00000' are region/HQ rows, not individual offices",
            "Rows with no P code are orphan metadata rows (skipped)",
            "Col H format: 'N. FullName' where N is sequential index (1-116)",
            "Col O (L/M/S/XS) is the office size classification",
            "#N/A and '0' values in Excel are treated as empty/null",
            "Single region กฟฉ.1 in current data; model supports multiple",
        ],
    }

    if not EXCEL_PATH.exists():
        report["errors"].append(f"Excel file not found: {EXCEL_PATH}")
        return None, report

    wb = openpyxl.load_workbook(str(EXCEL_PATH))
    ws = wb["Sheet1"]

    # ─── Read all rows ───
    raw_rows = []
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        row = {}
        for cell in r:
            row[cell.column_letter] = cell.value
        raw_rows.append(row)

    print(f"Read {len(raw_rows)} data rows from Sheet1")

    # ─── Deduplicate: take first occurrence of each P code ───
    seen_codes = set()
    office_rows = []
    region_rows = []
    skipped_rows = []
    duplicate_rows = 0

    for i, row in enumerate(raw_rows, start=2):
        p_code = safe_str(row.get("P"))
        if not p_code:
            skipped_rows.append({"row": i, "reason": "no P code"})
            continue
        if p_code == "D00000":
            region_rows.append(row)
            continue
        if p_code in seen_codes:
            duplicate_rows += 1
            continue  # Skip duplicate — keep only first occurrence
        seen_codes.add(p_code)
        office_rows.append(row)

    print(f"Unique offices: {len(office_rows)}")
    print(f"Duplicate rows skipped: {duplicate_rows}")
    print(f"Region rows: {len(region_rows)}, Orphan rows: {len(skipped_rows)}")

    report["stats"]["total_raw_rows"] = len(raw_rows)
    report["stats"]["unique_offices"] = len(office_rows)
    report["stats"]["duplicate_rows_skipped"] = duplicate_rows
    report["stats"]["region_rows"] = len(region_rows)
    report["stats"]["orphan_rows"] = len(skipped_rows)

    # ─── Region ───
    region = {
        "id": "R01",
        "code": "D00000",
        "name": "การไฟฟ้าส่วนภูมิภาค เขต 1 (ภาคตะวันออกเฉียงเหนือ)",
        "shortName": "กฟฉ.1",
    }

    # ─── Build Group8 (8 กฟจ.) ───
    # Collect unique parent8 values from deduplicated offices
    group8_map = {}
    for row in office_rows:
        name = safe_str(row.get("D"))
        short = safe_str(row.get("E"))
        if name and name not in group8_map:
            group8_map[name] = short
        elif name and short and not group8_map.get(name):
            group8_map[name] = short

    group8_list = []
    for i, name in enumerate(sorted(group8_map.keys()), 1):
        group8_list.append({
            "id": f"G8_{i:02d}",
            "name": name,
            "shortName": group8_map[name],
            "regionId": region["id"],
        })

    g8_id_by_name = {g["name"]: g["id"] for g in group8_list}
    print(f"Group8 (8 กฟจ.): {len(group8_list)}")

    # ─── Build Group17 (17 จุดรวมงาน) ───
    # Unique by (parent17 name col B). Use col G for short name.
    # The groupCode (col Q) in the first section maps 1:1 to these groups.
    group17_map = {}  # name -> {shortName, parent8Name}
    for row in office_rows:
        name = safe_str(row.get("B"))
        short = safe_str(row.get("G"))
        p8_name = safe_str(row.get("D"))
        group_code = safe_str(row.get("Q"))
        if name and name not in group17_map:
            group17_map[name] = {
                "shortName": short,
                "parent8Name": p8_name,
                "groupCode": group_code,
            }

    group17_list = []
    for name in sorted(group17_map.keys()):
        info = group17_map[name]
        p8_id = g8_id_by_name.get(info["parent8Name"], "")
        group17_list.append({
            "id": f"G17_{info['groupCode']}" if info["groupCode"] else f"G17_{name}",
            "code": info["groupCode"],
            "name": name,
            "shortName": info["shortName"],
            "parent8Id": p8_id,
            "regionId": region["id"],
        })
        if not p8_id:
            report["warnings"].append(f"Group17 '{name}' has no matching Group8")

    g17_id_by_name = {g["name"]: g["id"] for g in group17_list}
    print(f"Group17 (17 จุดรวมงาน): {len(group17_list)}")

    # ─── Build Group43 (43 กฟฟ.) ───
    # Unique by (parent43 name col C) within each parent17 context
    group43_map = {}  # (parent17_name, name) -> shortName
    g43_shortnames = {}  # col A only shows on "head office" rows
    for row in office_rows:
        c_name = safe_str(row.get("C"))
        b_name = safe_str(row.get("B"))
        a_short = safe_str(row.get("A"))
        if not c_name:
            continue
        key = (b_name, c_name)
        if key not in group43_map:
            group43_map[key] = {
                "parent17Name": b_name,
                "parent8Name": safe_str(row.get("D")),
            }
        if a_short and c_name:
            g43_shortnames[c_name] = a_short

    group43_list = []
    g43_counter = defaultdict(int)
    for key in sorted(group43_map.keys()):
        parent17_name, name = key
        info = group43_map[key]
        p17_id = g17_id_by_name.get(info["parent17Name"], "")
        p8_id = g8_id_by_name.get(info["parent8Name"], "")

        # Build a stable ID from parent17 code + sequence
        p17_code = ""
        for g17 in group17_list:
            if g17["name"] == info["parent17Name"]:
                p17_code = g17["code"]
                break
        g43_counter[p17_code] += 1
        idx = g43_counter[p17_code]

        group43_list.append({
            "id": f"G43_{p17_code}_{idx:02d}" if p17_code else f"G43_UNK_{idx:02d}",
            "name": name,
            "shortName": g43_shortnames.get(name, ""),
            "parent17Id": p17_id,
            "parent8Id": p8_id,
            "regionId": region["id"],
        })

    g43_id_by_key = {}
    for g43 in group43_list:
        # Map by name for office lookup
        g43_id_by_key[g43["name"]] = g43["id"]

    print(f"Group43 (43 กฟฟ.): {len(group43_list)}")
    if len(group43_list) not in (43, 44):
        report["warnings"].append(
            f"Expected ~43 Group43 entries but found {len(group43_list)}"
        )

    # ─── Build Offices ───
    offices = []
    for row in office_rows:
        p_code = safe_str(row.get("P"))
        h_val = safe_str(row.get("H"))
        idx, full_name = parse_index_and_name(h_val)
        short_name = safe_str(row.get("I"))
        size = safe_str(row.get("O"))
        parent43_name = safe_str(row.get("C"))
        parent17_name = safe_str(row.get("B"))
        parent8_name = safe_str(row.get("D"))
        shortname8 = safe_str(row.get("E"))
        shortname17 = safe_str(row.get("G"))
        group_code = safe_str(row.get("Q"))
        group_num = safe_int(row.get("R"))
        warehouse = safe_str(row.get("M"))
        biz_type = safe_str(row.get("N"))

        # Collect aliases
        alias_candidates = [
            safe_str(row.get("A")),
            safe_str(row.get("J")),
            safe_str(row.get("K")),
            safe_str(row.get("L")),
        ]
        aliases = sorted(set(a for a in alias_candidates if a and a != short_name))

        # Validate
        if not size or size not in ("L", "M", "S", "XS"):
            report["warnings"].append(f"Office {p_code} skipped: invalid/missing size '{size}'")
            continue

        # Resolve hierarchy IDs
        parent43_id = g43_id_by_key.get(parent43_name, "")
        parent17_id = g17_id_by_name.get(parent17_name, "")
        parent8_id = g8_id_by_name.get(parent8_name, "")

        office = {
            "code": p_code,
            "shortName": short_name,
            "fullName": full_name or short_name,
            "size": size,
            "index": idx,
            "parent43Id": parent43_id,
            "parent43Name": parent43_name,
            "parent17Id": parent17_id,
            "parent17Name": parent17_name,
            "parent8Id": parent8_id,
            "parent8Name": parent8_name,
            "regionId": region["id"],
            "regionName": region["shortName"],
            "shortName8": shortname8,
            "shortName17": shortname17,
            "groupCode": group_code,
            "groupNum": group_num,
            "aliases": aliases,
            "warehouseGroup": warehouse,
            "businessTypeCode": biz_type,
            "searchTokens": [],
        }
        office["searchTokens"] = generate_search_tokens(office)
        offices.append(office)

    print(f"Offices extracted: {len(offices)}")

    # ─── Stats ───
    size_counts = Counter(o["size"] for o in offices)
    report["stats"]["offices_by_size"] = dict(sorted(size_counts.items()))
    report["stats"]["total_offices"] = len(offices)
    report["stats"]["group8_count"] = len(group8_list)
    report["stats"]["group17_count"] = len(group17_list)

    # ─── Data Validation ───
    import hashlib
    codes_list = sorted(o["code"] for o in offices)
    data_hash = hashlib.sha256(json.dumps(codes_list).encode()).hexdigest()[:12]
    report["validation"] = {
        "totalOffices": len(offices),
        "uniquePCodes": len(seen_codes),
        "dataHash": data_hash,
        "allSizesValid": all(o["size"] in ("L", "M", "S", "XS") for o in offices),
        "allCodesNonEmpty": all(o["code"] for o in offices),
        "status": "OK" if len(offices) > 0 else "ERROR",
    }
    report["stats"]["group43_count"] = len(group43_list)

    # ─── Validate hierarchy consistency ───
    for o in offices:
        if not o["parent43Id"]:
            report["warnings"].append(f"Office {o['code']} ({o['shortName']}) has no Group43 match for '{o['parent43Name']}'")
        if not o["parent8Id"]:
            report["warnings"].append(f"Office {o['code']} ({o['shortName']}) has no Group8 match")

    # ─── Build Hierarchy Tree ───
    hierarchy = {
        "region": {**region},
        "children": [],
    }

    for g8 in group8_list:
        g8_node = {"id": g8["id"], "name": g8["name"], "shortName": g8["shortName"], "type": "group8", "children": []}
        for g17 in group17_list:
            if g17["parent8Id"] == g8["id"]:
                g17_node = {"id": g17["id"], "code": g17["code"], "name": g17["name"], "shortName": g17["shortName"], "type": "group17", "children": []}
                for g43 in group43_list:
                    if g43["parent17Id"] == g17["id"]:
                        g43_node = {"id": g43["id"], "name": g43["name"], "shortName": g43["shortName"], "type": "group43", "children": []}
                        for o in offices:
                            if o["parent43Id"] == g43["id"]:
                                g43_node["children"].append({
                                    "code": o["code"],
                                    "shortName": o["shortName"],
                                    "fullName": o["fullName"],
                                    "size": o["size"],
                                    "type": "office",
                                })
                        g17_node["children"].append(g43_node)
                g8_node["children"].append(g17_node)
        hierarchy["children"].append(g8_node)

    # ─── Assemble master output ───
    master = {
        "_meta": {
            "version": "1.0.0",
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "sourceFile": "Peaname.xlsx",
            "description": "PEA office master data — กฟฉ.1 (การไฟฟ้าส่วนภูมิภาค เขต 1)",
            "totalOffices": len(offices),
            "dataHash": report["validation"]["dataHash"],
        },
        "region": region,
        "groups8": group8_list,
        "groups17": group17_list,
        "groups43": group43_list,
        "offices": offices,
        "hierarchy": hierarchy,
    }

    return master, report


def main():
    print("=" * 60)
    print("PEA Data Extraction & Normalization")
    print("=" * 60)

    DATA_DIR.mkdir(parents=True, exist_ok=True)

    global EXCEL_PATH
    if not EXCEL_PATH.exists():
        alt_path = PROJECT_ROOT / "Peaname.xlsx"
        if alt_path.exists():
            EXCEL_PATH = alt_path
        else:
            print(f"ERROR: Cannot find Peaname.xlsx")
            sys.exit(1)

    master, report = extract()

    if master is None:
        print("\n[FAIL] Extraction failed.")
        with open(REPORT_PATH, "w", encoding="utf-8") as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        sys.exit(1)

    # Print summary
    if report["errors"]:
        print("\n[ERROR] ERRORS:")
        for e in report["errors"]:
            print(f"  - {e}")

    warn_count = len(report["warnings"])
    if warn_count:
        print(f"\n[WARN] WARNINGS ({warn_count}):")
        for w in report["warnings"][:5]:
            print(f"  - {w}")
        if warn_count > 5:
            print(f"  ... and {warn_count - 5} more (see validation_report.json)")

    print(f"\n[STATS]:")
    for k, v in report["stats"].items():
        print(f"  {k}: {v}")

    # Write outputs
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(master, f, ensure_ascii=False, indent=2)
    print(f"\n[OK] Master data: {OUTPUT_PATH} ({OUTPUT_PATH.stat().st_size:,} bytes)")

    with open(REPORT_PATH, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"[OK] Report: {REPORT_PATH}")

    if report["errors"]:
        print("\n[FAIL] Completed with ERRORS.")
        sys.exit(1)
    else:
        print("\n[OK] Extraction completed successfully!")


if __name__ == "__main__":
    main()
